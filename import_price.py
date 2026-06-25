from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Нужен пакет openpyxl: pip install openpyxl") from exc


PRICE_FILE_NAME = "прайс_русский_лес_25_06_26.xlsx"
SHEET_NAME = "Лист1"
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "data" / "products.json"


TRANSLIT = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "c",
        "ч": "ch",
        "ш": "sh",
        "щ": "shch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = text.replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", text)


def normalize_for_compare(value: str) -> str:
    return unicodedata.normalize("NFC", value).casefold()


def find_price_file() -> Path:
    expected = normalize_for_compare(PRICE_FILE_NAME)
    for path in BASE_DIR.glob("*.xlsx"):
        if normalize_for_compare(path.name) == expected:
            return path

    candidates = list(BASE_DIR.glob("*13.05.2026*.xlsx"))
    if candidates:
        return candidates[0]

    raise FileNotFoundError(f"Не найден файл {PRICE_FILE_NAME}")


def parse_price(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and value > 0:
        return int(round(value))

    text = clean_text(value).lower()
    if not text:
        return None
    if any(word in text for word in ["цена", "ед изм", "ед.изм", "толщина", "ширина", "длина", "кол-во", "размер"]):
        return None

    match = re.search(r"\d+(?:[\s.,]\d+)*", text)
    if not match:
        return None

    number = match.group(0).replace(" ", "").replace(",", ".")
    try:
        price = float(number)
    except ValueError:
        return None

    if price <= 0:
        return None

    return int(round(price))


def format_number(value: float) -> str:
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def normalize_dimension_text(value) -> str:
    text = clean_text(value)
    text = text.replace(",", ".")
    text = re.sub(r"\s*-\s*", "-", text)
    return text


def format_mm(value) -> str:
    text = normalize_dimension_text(value)
    if not text:
        return ""
    if re.search(r"[а-яa-z]", text.lower()):
        return text

    return f"{text} мм"


def format_length(value) -> str:
    text = normalize_dimension_text(value)
    if not text:
        return ""

    has_unit = bool(re.search(r"[а-яa-z]", text.lower()))

    def replace_number(match: re.Match[str]) -> str:
        raw = match.group(0).replace(",", ".")
        try:
            number = float(raw)
        except ValueError:
            return raw
        if number > 100:
            number = number / 1000
        return format_number(number)

    formatted = re.sub(r"\d+(?:[.,]\d+)?", replace_number, text)
    if has_unit:
        return formatted.replace("м.", "м")
    return f"{formatted} м"


def normalize_unit(value) -> str:
    text = clean_text(value).lower()
    compact = text.replace(".", "").replace(" ", "")

    if compact in {"м2", "м²", "m2"}:
        return "м²"
    if compact in {"пм", "п/м", "погм"}:
        return "п/м"
    if compact in {"шт", "штука", "штук"}:
        return "шт"
    if compact in {"м3", "м³", "m3"}:
        return "м³"

    return text or "шт"


def detect_category(name: str, section: str = "") -> str:
    haystack = f"{name} {section}".lower().replace("ё", "е")

    if "osb" in haystack or "осб" in haystack:
        return "OSB"
    if "мебельный щит" in haystack or " щит" in haystack or haystack.startswith("щит"):
        return "мебельный щит"
    if "имитац" in haystack:
        return "имитация бруса"
    if "блок хаус" in haystack or "блок-хаус" in haystack:
        return "блок-хаус"
    if "террас" in haystack:
        return "террасная доска"
    if "доска пола" in haystack or "половой шпунт" in haystack:
        return "доска пола"
    if "вагонка" in haystack:
        return "вагонка"
    if "планкен" in haystack:
        return "планкен"
    if "полок" in haystack:
        return "полок для бани"
    if "брусок" in haystack:
        return "брусок"
    if "клеен" in haystack and "брус" in haystack:
        return "клееный брус"
    if "брус" in haystack:
        return "брус"
    if "фанера" in haystack or "фсф" in haystack or "фк" in haystack:
        return "фанера"
    if "погонаж" in haystack:
        return "погонаж"
    if any(word in haystack for word in ["наличник", "уголок", "галтель", "плинтус", "нащельник", "шпатик"]):
        return "погонаж"
    if any(word in haystack for word in ["доска", "дощечка", "пиломатериал"]):
        return "доска"
    if any(word in haystack for word in ["ступень", "столешница"]):
        return "изделия из дерева"
    if any(word in haystack for word in ["акватекс", "покрытие", "масло", "лак"]):
        return "покрытия"

    return "прочее"


def detect_wood_type(name: str, section: str = "") -> str:
    haystack = f"{name} {section}".lower().replace("ё", "е")

    if "osb" in haystack or "осб" in haystack:
        return "OSB"
    if "ламинирован" in haystack and "фанер" in haystack:
        return "ламинированная фанера"
    if "фсф" in haystack:
        return "фанера ФСФ"
    if "фк" in haystack:
        return "фанера ФК"
    if "сосна" in haystack:
        return "сосна"
    if any(word in haystack for word in ["листвен", "листв", "лист-ца", "лист-ница"]):
        return "лиственница"
    if "липа" in haystack:
        return "липа"
    if "осина" in haystack:
        return "осина"
    if "кедр" in haystack:
        return "кедр"
    if "берез" in haystack:
        return "берёза"

    return ""


def detect_grade(name: str) -> str:
    text = clean_text(name).upper().replace("Ё", "Е")
    tokens = re.findall(r"[A-ZА-Я0-9/]+", text)

    for grade in ("1/2", "2/2", "2/3", "3/3"):
        if grade in tokens or grade in text:
            return grade

    if "ЭКСТРА" in tokens or "ЭКСТРА" in text or "Э" in tokens:
        return "Экстра"

    if any(token in {"AB", "АВ", "ABC", "АВС"} for token in tokens):
        return "AB"
    if any(token in {"BC", "ВС"} for token in tokens):
        return "BC"
    if any(token in {"A", "А"} for token in tokens):
        return "A"
    if any(token in {"B", "В"} for token in tokens):
        return "B"
    if any(token in {"C", "С"} for token in tokens):
        return "C"

    return ""


def determine_purpose(category: str) -> list[str]:
    if category == "вагонка":
        return ["для стен", "для потолка", "для бани"]
    if category == "планкен":
        return ["для фасада"]
    if category == "террасная доска":
        return ["для террасы"]
    if category in {"брус", "клееный брус", "брусок"}:
        return ["для каркаса", "для обрешётки"]
    if category in {"фанера", "OSB"}:
        return ["для черновых работ", "для опалубки", "для каркаса"]
    if category == "мебельный щит":
        return ["для мебели", "для столешницы"]
    if category == "доска пола":
        return ["для пола"]
    if category == "полок для бани":
        return ["для бани"]
    if category == "погонаж":
        return ["для отделки"]

    return ["для строительства", "для отделки"]


def build_description(category: str, wood_type: str) -> str:
    wood_forms = {
        "сосна": "сосны",
        "лиственница": "лиственницы",
        "липа": "липы",
        "осина": "осины",
        "кедр": "кедра",
        "берёза": "берёзы",
    }
    wood = ""
    if wood_type in wood_forms:
        wood = f" из {wood_forms[wood_type]}"
    descriptions = {
        "вагонка": f"Деревянная вагонка{wood} для внутренней отделки стен, потолка и бани.",
        "имитация бруса": f"Имитация бруса{wood} для отделки стен и создания фактуры деревянного дома.",
        "блок-хаус": f"Блок-хаус{wood} для фасадной и интерьерной отделки.",
        "планкен": f"Планкен{wood} для фасада, ограждений и декоративной отделки.",
        "террасная доска": f"Террасная доска{wood} для настилов, крыльца и открытых зон.",
        "брусок": f"Строганный брусок{wood} для каркаса, обрешётки и столярных работ.",
        "клееный брус": f"Строганный клееный брус{wood} для конструкций, столярных работ и интерьера.",
        "брус": f"Брус{wood} для каркасов, стоек, балок и строительных конструкций.",
        "фанера": "Листовая фанера для черновых работ, опалубки, мебели и строительных задач.",
        "OSB": "OSB-плита для пола, стен, кровли, обшивки и каркасных работ.",
        "мебельный щит": "Мебельный щит для мебели, столешниц, ступеней и столярных изделий.",
        "полок для бани": f"Полок{wood} для бани, сауны и парной.",
        "погонаж": f"Погонаж{wood} для аккуратной отделки стыков, углов и проёмов.",
        "доска пола": f"Доска пола{wood} для устройства деревянного пола.",
        "доска": f"Доска{wood} для строительства, отделки и черновых работ.",
        "изделия из дерева": "Готовые элементы из дерева для дома, участка и интерьера.",
        "покрытия": "Покрытие для защиты и декоративной обработки древесины.",
    }
    return descriptions.get(category, "Пиломатериал со склада Русский Лес в Алматы.")


def detect_availability(note) -> str:
    text = clean_text(note).lower().replace("ё", "е")
    if re.search(r"\d+\s*(?:шт|штук)|по\s*\d+|остат|ост\b|не много|немного|мало", text):
        return "в наличии"
    return "уточнить"


def image_for_product(name: str, category: str, wood_type: str = "") -> str:
    category_images = {
        "блок-хаус": "images/blok-haus.jpg",
        "имитация бруса": "images/imitaciya-brusa.jpg",
        "террасная доска": "images/terrasnaya-doska.jpg",
    }
    if category == "вагонка":
        if wood_type == "липа":
            return "images/lipa_vagonka.jpg"
        if "евро" in name.lower():
            return "images/evro_vagonka.jpg"
        return "images/vagonka.PNG"
    if category == "OSB":
        return "images/OSB.PNG"
    if category == "фанера":
        return "images/fanera.PNG"
    if category == "клееный брус":
        return "images/kleenyy-brus.png"
    if category == "планкен" and wood_type == "лиственница":
        return "images/planken_listvinica.jpg"
    if category in category_images:
        return category_images[category]
    if category in {"брус", "клееный брус", "брусок", "доска"}:
        return "images/Brusok-sosna.jpg"
    return "images/hero.png"


def clean_name(value) -> str:
    text = clean_text(value)
    text = text.replace("(", " ").replace(")", " ")
    text = text.replace('"', " ").replace("«", " ").replace("»", " ")
    text = re.sub(r"\s*-\s*", " ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    return re.sub(r"\s+", " ", text).strip()


def slugify(value: str) -> str:
    text = value.lower().replace("×", "x").replace("*", "x")
    text = text.translate(TRANSLIT)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "product"


def parse_dimensions_from_name(name: str) -> tuple[str, str, str]:
    match = re.search(
        r"(\d+(?:[.,]\d+)?)\s*[*xх×]\s*(\d+(?:[.,]\d+)?)\s*[*xх×]\s*(\d+(?:[.,]\d+)?)",
        name,
        re.IGNORECASE,
    )
    if not match:
        return "", "", ""
    thickness, width, length = match.groups()
    return format_mm(thickness), format_mm(width), format_length(length)


def parse_countertop_sizes(size_text: str) -> tuple[str, str, str]:
    text = normalize_dimension_text(size_text).lower()

    thickness = ""
    width = ""
    length = ""

    thickness_match = re.search(r"(\d+(?:[.,]\d+)?)\s*мм\s*тол", text)
    if thickness_match:
        thickness = format_mm(thickness_match.group(1))

    width_match = re.search(r"(\d+(?:[.,]\d+)?)\s*мм\s*шир", text)
    if width_match:
        width = format_mm(width_match.group(1))

    length_match = re.search(r"(\d+(?:[.,]\d+)?)\s*мм\s*длин", text)
    if length_match:
        length = format_length(length_match.group(1))

    return thickness, width, length


def build_display_fields(name: str, thickness: str, width: str, length: str) -> tuple[str, str, str]:
    dimension_label = " × ".join(part for part in [thickness, width, length] if part)
    catalog_name = name

    if dimension_label:
        catalog_name = re.sub(
            r"\d+(?:[.,]\d+)?\s*[*xх×]\s*\d+(?:[.,]\d+)?\s*[*xх×]\s*\d+(?:[.,]\d+)?",
            "",
            name,
            count=1,
            flags=re.IGNORECASE,
        )
        catalog_name = re.sub(r"\s+", " ", catalog_name).strip(" ,-") or name

    display_name = f"{catalog_name} — {dimension_label}" if dimension_label else catalog_name
    return catalog_name, dimension_label, display_name


def row_has_header_words(row_values: list) -> bool:
    values = [clean_text(value).lower() for value in row_values if value is not None]
    if not values:
        return False

    if values[0] == "наименование":
        return True

    columns = " ".join(values[1:])
    return "цена" in columns or "ед изм" in columns or "ед.изм" in columns


def is_service_row(name: str) -> bool:
    lowered = name.lower()
    return any(word in lowered for word in ["компания", "г.алматы", "тел/факс", "наименование"])


def mode_from_section(section: str) -> str:
    lowered = section.lower().replace("ё", "е")
    if "пиломатериал обрез" in lowered:
        return "lumber"
    if "столешниц" in lowered:
        return "countertop"
    return "standard"


def product_from_row(row_values: list, row_number: int, section: str, mode: str) -> dict | None:
    name = clean_name(row_values[0])
    if not name or is_service_row(name):
        return None

    if mode == "lumber":
        price = parse_price(row_values[3] if len(row_values) > 3 else None)
        if price is None:
            return None
        note = clean_text(row_values[4] if len(row_values) > 4 else "")
        thickness, width, length = parse_dimensions_from_name(name)
        name = f"Пиломатериал обрезной {name}"
        unit = "шт"
    elif mode == "countertop":
        price = parse_price(row_values[5] if len(row_values) > 5 else None)
        if price is None:
            return None
        size_text = clean_text(row_values[1] if len(row_values) > 1 else "")
        note = size_text
        unit = normalize_unit(row_values[4] if len(row_values) > 4 else "шт")
        thickness, width, length = parse_countertop_sizes(size_text)
    else:
        price = parse_price(row_values[5] if len(row_values) > 5 else None)
        if price is None:
            return None
        thickness = format_mm(row_values[1] if len(row_values) > 1 else "")
        width = format_mm(row_values[2] if len(row_values) > 2 else "")
        length = format_length(row_values[3] if len(row_values) > 3 else "")
        unit = normalize_unit(row_values[4] if len(row_values) > 4 else "")
        note = clean_text(row_values[6] if len(row_values) > 6 else "")

    category = detect_category(name, section)
    wood_type = detect_wood_type(name, section)
    grade = detect_grade(name)

    dimensions = " ".join(part for part in [thickness, width, length] if part)
    variant_label = " × ".join(part for part in [thickness, width, length] if part) or grade or "Стандарт"
    catalog_name, dimension_label, display_name = build_display_fields(name, thickness, width, length)
    slug_parts = [name, grade, dimensions]

    characteristics = {
        "Единица": unit,
    }
    if wood_type:
        characteristics["Порода"] = wood_type
    if grade:
        characteristics["Сорт"] = grade
    if thickness:
        characteristics["Толщина"] = thickness
    if width:
        characteristics["Ширина"] = width
    if length:
        characteristics["Длина"] = length

    return {
        "source_row": row_number,
        "slug_base": slugify(" ".join(slug_parts)),
        "name": name,
        "catalog_name": catalog_name,
        "dimension_label": dimension_label,
        "display_name": display_name,
        "category": category,
        "wood_type": wood_type,
        "grade": grade,
        "thickness": thickness,
        "width": width,
        "length": length,
        "purpose": determine_purpose(category),
        "availability": detect_availability(note),
        "description": build_description(category, wood_type),
        "image": image_for_product(name, category, wood_type),
        "base_price": price,
        "unit": unit,
        "note": note,
        "variant_label": variant_label,
        "variants": [],
        "characteristics": characteristics,
    }


def make_unique_slugs(products: list[dict]) -> None:
    used: Counter[str] = Counter()
    for product in products:
        base = product.pop("slug_base")
        used[base] += 1
        product["slug"] = base if used[base] == 1 else f"{base}-{used[base]}"


def mark_popular_products(products: list[dict]) -> None:
    preferred_categories = [
        "OSB",
        "фанера",
        "вагонка",
        "брусок",
        "клееный брус",
        "брус",
        "доска",
        "планкен",
        "террасная доска",
    ]
    marked = set()

    for product in products:
        product["popular"] = False

    for category in preferred_categories:
        for product in products:
            if product["category"] == category and category not in marked:
                product["popular"] = True
                marked.add(category)
                break


def import_products() -> tuple[list[dict], int, Counter]:
    workbook_path = find_price_file()
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    products: list[dict] = []
    skipped = 0
    current_section = ""
    current_mode = "standard"

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = list(row)
        name = clean_text(row_values[0] if row_values else "")

        if not name:
            continue

        if row_number <= 3 and is_service_row(name):
            continue

        has_header = row_has_header_words(row_values)
        if has_header:
            if not is_service_row(name) and name.lower() not in {"наименование"}:
                current_section = name
                current_mode = mode_from_section(current_section)
            else:
                current_section = ""
                current_mode = "standard"
            continue

        product = product_from_row(row_values, row_number, current_section, current_mode)

        if product:
            products.append(product)
            continue

        if not any(parse_price(value) for value in row_values[1:]):
            if not is_service_row(name) and name.lower() not in {"наименование"}:
                current_section = name
                current_mode = mode_from_section(current_section)
            if row_number > 3 and not is_service_row(name):
                skipped += 1
            continue

        if row_number > 3 and not is_service_row(name):
            skipped += 1

    for index, product in enumerate(products, start=1):
        product["id"] = index

    make_unique_slugs(products)
    mark_popular_products(products)

    categories = Counter(product["category"] for product in products)
    return products, skipped, categories


def main() -> None:
    products, skipped, categories = import_products()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Импортировано товаров: {len(products)}")
    print(f"Пропущено строк без цены: {skipped}")
    print("Товары по категориям:")
    for category, count in categories.most_common():
        print(f"- {category}: {count}")
    print(f"Файл сохранён: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
